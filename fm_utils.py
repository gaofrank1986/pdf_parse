#  import PyPDF2
from tabula import read_pdf
import json
import os
import logging
from collections import OrderedDict
import pandas as pd
import re
from sqlalchemy import exists,and_,or_
from openpyxl import Workbook


class FmUtils(object):

    def __init__(self):
        logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
        logging.disable(level=logging.INFO)
        self.mode = 1

    def read_summary(self,path,pages):

        df = read_pdf(path, pages = pages,silent=True, multiple_tables=True, pandas_option={'header':None})

        ww = pd.concat(df[:]).fillna(' ')
        ww.index = range(0,ww.shape[0])
        self._buf = ww.copy()

        self._preprocess_df_step1(ww)
        self._smr_preprocess_df(ww)
        self._buf2 = ww.copy()

        res = self._clean_df(ww)
        self._buf3 = res

        # 如果查出季度数据，去掉
        mark =[]
        for i in range(len(res)):
            if '营业收入' in res[i]:
                mark.append(i)
        if len(mark) > 1:
            e = res[:mark[1]]
        else:
            e = res

        self._buf4 = e
        ans = OrderedDict()
        for d in e:
            d = d.split()
            ans[d[0]] = (self._comma_sep_string_to_num(d[1]),self._comma_sep_string_to_num(d[2]))

        return(ans)

    def _smr_preprocess_df(self,df):
        # 第一列内容
        w = list(df.iloc[:,0])
        # 第一行内容
        v = list(df.iloc[0,:])
        # 第一列去括号。。
        for i in range(len(w)):
            tmp = [self._format_cell(x) for x in w[i].split()]
            w[i] = ''.join(tmp)

        res =[]
        pool = [ '归属','扣除非' ]
        pool2 = ['润','率','产','益']
        for i in range(len(w)):
            tmp = [x in w[i] for x in pool]
            if any(tmp):
                res.append(i)
        for i in res:
            for kk in range(1,3):
                if(i+kk>len(w)-1):
                    break
                if any([w[i].endswith(x) for x in pool2]):
                    break
                df.loc[i+kk,0] = w[i+kk].replace(' ','')
                df.loc[i,:] += df.loc[i+kk,:]
                # 加入下一行，结束判定可有上一个if完成
                w[i] += w[i+kk]
                for j in range(len(v)):
                    df.iloc[i+kk,j] = ''

    #------------------------------------------------------------ 
    #--------------------- read and clean part------------------
    #------------------------------------------------------------ 

    def _preprocess_df_step2(self,df):
        w = list(df.iloc[:,0])
        v = list(df.iloc[0,:])
        res =[]
        pool = [ '购建固定资产', '处置固定资产','销售商品、提','经营活动产生']
        pool2 =['金','净额']
        for i in range(len(w)):
            if any([x in w[i] for x in pool]):
                res.append(i)
        for i in res:
            for kk in range(1,3):
                if(i+kk>len(w)-1):
                    break
                if any([w[i].endswith(x) for x in pool2]):
                #  if w[i].endswith('金') or w[i].endswith('净额'):
                    break
                df.loc[i+kk,0] = w[i+kk].replace(' ','')
                df.loc[i,:] += df.loc[i+kk,:]
                # clear the line, if added to above line
                for j in range(len(v)):
                    df.iloc[i+kk,j] = ''
                #  if w[i+kk].endswith('金') or w[i+kk].endswith('净额'):
                if any([w[i+kk].endswith(x) for x in pool2]):
                    break



    def _preprocess_df_step1(self,df):
        #括号跨越
        w = list(df.iloc[:,0])
        v = list(df.iloc[0,:])
        for i in range(len(w)):
            if (df.iloc[i,0].count('(') == 1) and (df.iloc[i,0].count(')') == 0):
                for kk in range(1,3):
                    # 超出表格·
                    if(i+kk>len(w)-1):
                        break
                    df.loc[i+kk,0] = w[i+kk].replace(' ','')
                    df.loc[i,:] += df.loc[i+kk,:]
                    for j in range(len(v)):
                        df.iloc[i+kk,j] = ''
                    if w[i+kk].endswith(')') :
                        break

    def _preprocess_df_step3(self,df):
        #数字/空白
        #空白/数字
        w = list(df.iloc[:,0])
        v = list(df.iloc[0,:])
        for i in range(len(w)):
            flag = [self._is_str_valid(df.iloc[i,-2]),self._is_str_valid(df.iloc[i,-1])]
            if (flag[0] and df.iloc[i,-1] ==' '):
                df.iloc[i,-1] ='00000.00'
            elif(flag[1] and df.iloc[i,-2] == ' '):
                df.iloc[i,-2] ='00000.00'


    def read_and_clean(self, path, pages="",check_multi = True, mode = 1):
        # mode 0 ,忽略 multi_check
        # mode 1 ,正常
        # mode 2 , process_line 去除 [0] 和 最后两个数字之间的部分
        # mode 3 ,处理 2017数据合并,母公司，2016数据合并,数据母公A(默认忽略前处理step3)
        # mode 4, mode0+mode3(默认忽略前处理step3)
        # mode 5, 忽略前处理step3

        self.mode = mode

        if not(pages):
            pages = "all"
        df = read_pdf(path, pages = pages,silent=True, multiple_tables=True, pandas_option={'header':None})

        for k in range(len(df)):
            ncol = df[k].shape[1]
            # 如果在前2/3的部分发现大于6列则判定双列表
            if  ncol > 6 and k < round(len(df)*2/3) and mode!=3  and mode!=4:
                parallel_tab_flag = True
                #  raise Exception(path,ncol,"parallel table")

        df  =pd.concat(df[:]).fillna(' ')
        #  df  =pd.concat(df[:]).fillna('blank')
        df.index = range(0,df.shape[0])
        self._oldest = df.copy()

        #----------preprocessing--------------
        self._preprocess_df_step1(df)
        self._preprocess_df_step2(df)
        if mode <3:
            #  self._preprocess_df_step3(df)
            pass
        self._buf = df.copy()
        # 清理数据
        res = self._clean_df(df)
        # 标准化index
        res = self.normalize_table(res)
        self._buf2 = res.copy()

        # 处理子母表
        # a. 如果有重复
        #       起始为m[0]
        #       如果m[0]和[m1]相差1
        #       如果有m2,结束为m2,否则为下个pos的m0,或者表的最后一行
        #       
        new_output = []
        pos = self.mark_multi_table(res)
        self._buf3 = pos.copy()
        if check_multi and mode!=0 and mode!=4 :
            flag = (len(pos['t1'])==0) or (len(pos['t2'])==0) or (len(pos['t3'])==0)
            if flag:
                raise Exception("table missing")
            flag = (len(pos['t1'])>1) or (len(pos['t2'])>1) or (len(pos['t3'])>1)
            if flag:
                new_output=[]
                for i in range(1,4):
                    # 一般情况
                    # start = m[0]
                    # end = 下个表格其实/所有记录的最后一行位置
                    m = pos['t'+str(i)]
                    start = m[0]
                    if(i<2):
                        end = pos['t'+str(i+1)][0]
                    else:
                        end = len(res)-1

                    if len(m) > 1:
                        if (m[1]-m[0])==1:
                            if(len(m)>2):
                                end = m[2]
                        else:
                            end = m[1]

                    for j in range(start,end):
                        new_output.append(res[j])
                    pos['t'+str(i)] = {'start':start,'end':end}
            self._buf4 = pos

        if not(len(new_output)==0):
            res = new_output
        #--------------
        e,tmp = self._formatted_list_to_ordered_dic(res)

        return(e)


    def process_line(self,res):
        # 处理单行
        # 1. 对每行用_format_index进行预处理
        #    a. 少于两个数字，删除当前行
        #    b. 如果最后一个comp是文字，就去掉当前行
        #    c. 如果前两个都是文字，去掉第一个
        newline = [self._format_cell(i) for i in res.split()]
        flagline=[self._is_str_valid(i) for i in newline]

        if flagline.count(True) < 2:
            ans = ('')
        else:
            # abandon the line start with a valid number
            if flagline[0] == True:
                newline=['']
            #如果最后一个不是数字，删除
            elif flagline[-1] == False:
                newline.pop();
            #  elif (len(flagline[:-2]) > 1 and mode == 2):
                #  # 去掉第一个，和最后两个数字中间的
                #  for kk in range(len(flagline[:-2])-1):
                    #  # newline在删element过程中会变化，所以用len(flagline)
                    #  newline.remove(newline[len(flag[:-2])-1-kk])
            ans = ' '.join(newline)
        if (self.mode == 3 or self.mode == 4) and len(newline)==5:
            ans = ' '.join([newline[0],newline[1],newline[3]])
        return ans


    def normalize_table(self,output):
        for i in range(len(output)):
            tmp = output[i].split()
            if('销售商品、提供劳务收' in tmp[0]):
                tmp[0] = '销售商品、提供劳务收到的现金'
            if('经营活动产生的现' in tmp[0]):
                tmp[0] = '经营活动产生的现金流量净额'
            if tmp[0] in ['股东权益合计','所有者权益合计']:
                tmp[0] = '所有者权益'
            if  '购建固定资产' in tmp[0] :
                tmp[0] = '构建固定资产等所支付现金'
            if  '处置固定资产' in tmp[0] :
                tmp[0] = '处置固定资产等所收回现金净额'
            output[i]  = ' '.join(tmp)
        return(output)


    def mark_multi_table(self,output):
        # 通过下面的string标记合并报表和母公司报表的起始位置
        t1 = set(['货币资金'])
        t2 = set(['营业收入','营业总收入'])
        t3 = set(['销售商品、提供劳务收到的现金'])
        pos ={'t1':[],'t2':[],'t3':[]}
        for i in range(len(output)):
            tmp =set()
            tmp.add(output[i].split()[0])
            if len(tmp.intersection(t1)) != 0:
                pos['t1'].append(i)
            elif len(tmp.intersection(t2)) != 0:
                pos['t2'].append(i)
            elif len(tmp.intersection(t3)) != 0:
                pos['t3'].append(i)
        return(pos)


    def _is_str_valid(self,s):
        if not(isinstance(s,str)):
            s = str(s)
        # 去字符串开头结尾的括号
        s =  re.sub(r'^\(([0-9,-.]+)\)$', r'\1', s)
        # 去百分号
        s = s.replace('%','')
        if re.match(r'^[-]?[0-9,]*\d$',s) or re.match(r'^[-]?[0-9,]*\d[.]\d+$',s) :
            return True
        else:
            return False

    # 清理表格
    #    a.合并dataframe
    #    b.处理单行
    def _clean_df(self,df):
        #dataframe 变成 list
        #  for i in range(ww.shape[0]):
            #  out = ' '.join(list(ww.loc[i,:]))
            #  res.append(out)
        for i in range(df.shape[1]):
            if i == 0:
                tmp = df.loc[:,i]
            else:
                tmp = tmp+' '+df.loc[:,i]
        res = list(tmp)
        # 清理每行
        res = [ x for x in res if len(x.split())>2] 
        res = [ self.process_line(x) for x in res ]
        res = [ x for x in res if len(x.split()) > 2 ]

        return(res)


    # 转为有序dict
    def _formatted_list_to_ordered_dic(self,new_out):
        e = OrderedDict()
        # 记录重复次数
        count_d = dict()
        for i in new_out:
            tmp = i.split()
            flag = e.get(tmp[0],-1)
            if flag==-1:
                e[tmp[0]]=[self._comma_sep_string_to_num(tmp[1]),self._comma_sep_string_to_num(tmp[2])]
                count_d[tmp[0]] = 0
            else:
                count_d[tmp[0]] +=1
        flag = [x > 2 for x in count_d.values()]
        if any(flag):
            raise Exception("Too many duplicates")
            pass

        return (e,count_d)

    def _format_cell(self,s):
        assert(len(s.split())<2)
        if not(self._is_str_valid(s)):
            # 去掉汉字数字+[.] or [、]
            s = re.sub(r'[一二三四五六七八九十]+[、.]','',s)
            # 去掉 十一，十二
            s = re.sub(r'[二三四五六七八九十]{2}','',s)
            # 类似 七75
            s = re.sub(r'[二三四五六七八九十][^\u4E00-\u9FA5]+','',s)
            # [\u4E00-\u9FA5] 去掉类似(一) (二)
            #  s = re.sub(r'\([\u4E00-\u9FA5]{1,2}[、.]?\)','',s)
            # 如果不是数字项,去掉数字及.
            s = re.sub(r'[0-9.]','',s)
            #  s = s.replace('-','')
            if s == '-' or s == '—':
                s = '00000'
            # 去掉括号及其中内容
            s = re.sub(r'\(.*\)','',s)
            s = re.sub(r'\).*','',s)
            s = re.sub(r'\(.*','',s)
            # 去掉中文括号及其中内容
            s = re.sub(r'（.*）','',s)
            s = re.sub(r'）.*','',s)
            s = re.sub(r'（.*','',s)
            clist = ['其中:','减:','加:']
            for i in clist:
                s = s.replace(i,'')
        else:
            # 去除数字周围的括号
            s = re.sub(r'^\(([0-9-,.]+)\)$', r'\1', s)
            #如果是1位或者2位数字,删除
            s = re.sub(r'^\d\d?$','',s)
        if '注' in s:
            s = ''
        if '增加' in s:
            s = ''
        #    a. 去掉长度小于2的（一般是注释）(不然‘库存’‘商誉’)
        if len(s)<2:
            s = ''

        return(s)


    def _comma_sep_string_to_num(self,s):
        ##将逗号分隔的数字转成数字
        if not(isinstance(s,str)):
            raise Exception('no string input cannot be casted into float')
        if not(s):
            return(0)
        if not(self._is_str_valid(s)):
            raise Exception(s,"not valid number string")
        #  ss = s.replace(',','')
        #  ss = ss.replace('%','')
        ss = re.sub(r'[,%]','',s)

        sign = 1
        if (ss[0] == '-')and(ss.count('-')==1):
            sign = -1
            ss = ss.replace('-','')
        if ss.count('.') > 1:
            #  raise Exception(ss,'not valid format')
            print(ss)
            ss = '9999999999999999999'
        tmp = ss.replace('.','')
        if not(tmp.isdigit()):
            print(ss)
            ss = '9999999999999999999'
            #  raise Exception(ss,'string contain non number component')

        return(sign*float(ss))


    def swap_dict(self,idict):
        # best way to exchange key and value in a dict
        # res = dict((v,k) for k,v in a.items())
        return(dict((v,k) for k,v in idict.items()))

    def save_dict_to_json(self,path,idict,sort=False):

        with open(path, 'w') as fp:
            # add chinese output
            # add indentation
            #  json.dump(idict, fp,ensure_ascii=False,indent=4)
            json.dump(idict, fp,ensure_ascii=False,indent=4, sort_keys=sort)


    def load_from_json(self,path):
        data = json.load(open(path),object_pairs_hook=OrderedDict)
        return(data)

    def deal_with_duplicate_dict_values(self,idt):
        acc = 0
        for i in idt.keys():
            idt[i] = acc
            acc+=1


    def get_ratios(self,this,last):
        # ----------------------------------------------
        result = OrderedDict()
        prec = 2
        # total revenue
        result['revenue_incr'] = this['revenue']/last['revenue'] - 1
        result['gross_margin'] = 1 - this['op_cost']/this['revenue']
        result['total_expense_rate'] = (this['admin_expense']+this['sale_expense']+\
                this['finance_expense'])/this['revenue']
        result['sale_expense_rate'] = this['sale_expense']/this['revenue']
        result['admin_expense_rate'] = this['admin_expense']/this['revenue']
        result['finance_expense_rate'] = this['finance_expense']/this['revenue']
        result['reduced_net_profit_incr'] = this['op_profit']/last['op_profit'] - 1
        result['asset_debt_ratio'] = this['total_debt']/this['total_asset']
        result['acc_rcv_over_revenue'] = (this['acc_rcv']+this['note_rcv'])/this['revenue']
        result['net_working_cap'] = this['note_rcv']+ this['acc_rcv']+this['pre_paid']+this['other_rcv']\
                +this['inventory']-this['note_pyb']-this['acc_pyb']-this['pre_rcv']
        result['fixed_asset_ratio'] = this['fixed_asset']/this['total_asset']
        #new
        result['fixed_asset_incr'] = this['fixed_asset']/last['fixed_asset'] - 1
        result['on_building_over_fixed'] = this['on_building']/this['fixed_asset']
        result['cash_eq_over_total_asset'] = this['cash_and_eq']/this['total_asset']
        result['ROE'] = this['net_profit']/(0.5*this['equity']+0.5*last['equity'])
        result['net_profit_margin'] = this['net_profit']/this['revenue']
        result['total_turn_over'] = this['revenue']/(0.5*this['total_asset']+0.5*last['total_asset'])
        result['finance_leverage'] = this['total_asset']/(this['total_asset']-this['total_debt'])
        result['total_asset_incr'] = this['total_asset']/last['total_asset'] - 1
        # new
        result['equity_incr'] = this['equity']/last['equity'] - 1
        result['op_cash_over_net_profit'] = this['net_cash_from_op']/this['net_profit']
        result['capex_over_net_profit'] = this['capital_expenditure']/this['net_profit']
        result['smr_net_profit_incr'] = this['smr_deducted_net_profit']/last['smr_deducted_net_profit'] - 1
        result['smr_avg_roe'] = this['smr_avg_roe']

        for i in result.keys():
            result[i] = round(result[i]*100,prec)
        result['net_working_cap'] = round(result['net_working_cap']/10**8/100,prec)
        result['finance_leverage'] = round(result['finance_leverage']/100,prec)
        result['total_turn_over'] = round(result['total_turn_over']/100,prec)
        result['smr_avg_roe'] = round(result['smr_avg_roe']/100,prec)

        return(result)


    def save_to_excel(self,path,d):
        # '#,##0.00'
        # '#,##0.00$'
        # '$#,##0.00'

        wb = Workbook()
        ws = wb.active
        for i in d:
            ws.append([i,d[i][0],d[i][1]])
        for i in ws['b']+ws['c']:
            i.number_format = '#,##0.00'
        for i in ['a','b','c']:
            ws.column_dimensions[i].width = 40
        wb.save(path)


    def save_to_database(self,db_session,db,stock_id,year,d):
        keys = [str(stock_id)+'_'+str(year), str(stock_id)+'_'+str(int(year)-1)]
        name = self.load_from_json("./json/table_trans.json")
        for a_key in keys:
            (res,) =db_session.query(exists().where(db.fm_id==a_key))
            if not(res[0]):
                db_session.add(db(stock_id = stock_id,date = a_key[-4:]))
                db_session.commit()

        for i in d.keys():
            tmp = name.get(i,-99)
            if tmp!="0" and tmp!=-99:
                db_session.query(db).filter(db.fm_id == keys[0]).update({name[i]: d[i][0]})
                db_session.query(db).filter(db.fm_id == keys[1]).update({name[i]: d[i][1]})
            elif tmp==-99:
                print(i)

        db_session.commit()
